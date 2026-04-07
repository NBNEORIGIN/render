"""Render — NBNE sign product generator."""
import os
import json
from pathlib import Path
from datetime import datetime

from flask import Flask, render_template, jsonify, request, Response, send_file, send_from_directory, make_response, session, redirect, url_for
from flask_compress import Compress
from dotenv import load_dotenv

load_dotenv()

from config import SECRET_KEY, COLORS, BRAND_NAME, IMAGES_DIR
from models import init_db, get_db, release_db, dict_cursor, Product, Blank, User, SalesImport, SalesData
from jobs import submit_job, get_job, get_all_jobs, job_to_dict, start_workers

app = Flask(__name__)
app.secret_key = SECRET_KEY

# ── Authentication ─────────────────────────────────────────────────────────────
_PUBLIC_PATHS = {"/health", "/favicon.ico"}
_PUBLIC_PREFIXES = ("/login", "/static/", "/images/", "/etsy/oauth/")

LOGIN_HTML = """<!DOCTYPE html>
<html><head><title>Render — Login</title>
<style>
body{font-family:Arial,sans-serif;display:flex;align-items:center;justify-content:center;
     height:100vh;margin:0;background:#1a1a2e}
.box{background:#fff;padding:40px;border-radius:8px;width:320px;text-align:center}
h2{margin:0 0 24px}
label{display:block;text-align:left;font-size:13px;color:#555;margin-bottom:4px}
input{width:100%;padding:10px;margin-bottom:16px;box-sizing:border-box;
      border:1px solid #ccc;border-radius:4px;font-size:15px}
button{width:100%;padding:10px;background:#2563eb;color:#fff;border:none;
       border-radius:4px;cursor:pointer;font-size:16px}
button:hover{background:#1d4ed8}
.err{color:red;margin-top:12px;font-size:14px}
</style></head>
<body><div class="box">
<h2>Render</h2>
<form method="POST" action="/login">
  <label for="email">Email</label>
  <input type="email" id="email" name="email" required autofocus placeholder="you@nbnesigns.com">
  <label for="password">Password</label>
  <input type="password" id="password" name="password" required>
  <button type="submit">Sign in</button>
  __ERROR__
</form>
</div></body></html>"""

@app.before_request
def require_auth():
    """Protect all routes. Exempt: /health, /login, /static/, /images/."""
    if request.path in _PUBLIC_PATHS:
        return None
    if any(request.path.startswith(p) for p in _PUBLIC_PREFIXES):
        return None
    if not session.get("user_email"):
        if request.path.startswith("/api/"):
            return jsonify({"error": "Unauthorized"}), 401
        return redirect(url_for("login_page"))

@app.route("/login", methods=["GET"])
def login_page():
    if session.get("user_email"):
        return redirect("/")
    return LOGIN_HTML.replace('__ERROR__',"")

@app.route("/login", methods=["POST"])
def login_post():
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")
    user = User.authenticate(email, password)
    if user:
        session["user_email"] = user["email"]
        session["user_name"] = user["name"]
        return redirect("/")
    return LOGIN_HTML.replace('__ERROR__','<p class="err">Invalid email or password.</p>'), 401

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/health")
def health():
    return jsonify({"status": "ok"})

# ── App setup ──────────────────────────────────────────────────────────────────
Compress(app)
init_db()
start_workers()
IMAGES_DIR.mkdir(parents=True, exist_ok=True)

# HTML Template
# HTML template extracted to templates/index.html


@app.route('/')
def index():
    return render_template("index.html")


@app.route('/api/products', methods=['GET'])
def get_products():
    return jsonify(Product.all())


@app.route('/api/products', methods=['POST'])
def create_product():
    data = request.json
    Product.create(data)
    return jsonify({"success": True})


@app.route('/api/products/<m_number>', methods=['GET'])
def get_product(m_number):
    product = Product.get(m_number)
    if product:
        return jsonify(product)
    return jsonify({"error": "Not found"}), 404


@app.route('/api/products/<m_number>', methods=['PATCH'])
def update_product(m_number):
    data = request.json
    print(f"PATCH {m_number}: {data}")  # Debug logging

    # Check if this PATCH is changing qa_status to approved
    old_product = Product.get(m_number)
    was_approved = old_product and old_product.get('qa_status') == 'approved'

    Product.update(m_number, data)

    # Auto-publish on QA approval
    if (data.get('qa_status') == 'approved' and not was_approved):
        _trigger_auto_publish(m_number)

    return jsonify({"success": True})


@app.route('/api/products/<m_number>', methods=['DELETE'])
def delete_product(m_number):
    Product.delete(m_number)
    return jsonify({"success": True})


@app.route('/api/products/clear', methods=['DELETE'])
def clear_all_products():
    """Delete all products."""
    Product.clear_all()
    return jsonify({"success": True})


@app.route('/api/chat', methods=['POST'])
def chat_with_assistant():
    """Chat with AI assistant using OpenAI API."""
    import openai
    from config import OPENAI_API_KEY
    
    if not OPENAI_API_KEY:
        return jsonify({"error": "OpenAI API key not configured"}), 500
    
    data = request.json
    system_prompt = data.get('system_prompt', 'You are a helpful product development assistant.')
    messages = data.get('messages', [])
    
    try:
        client = openai.OpenAI(api_key=OPENAI_API_KEY)
        
        # Build messages with system prompt
        api_messages = [{"role": "system", "content": system_prompt}]
        api_messages.extend(messages)
        
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=api_messages,
            max_tokens=1000,
            temperature=0.7
        )
        
        assistant_response = response.choices[0].message.content
        return jsonify({"response": assistant_response})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/templates/csv')
def download_csv_template():
    """Download a CSV template for product data."""
    csv_content = '''m_number,description,size,color,ean,icon_files,orientation
M1001,Example Sign,saville,silver,5060000000001,icon.svg,landscape
M1002,Example Sign,saville,gold,5060000000002,icon.svg,landscape
M1003,Example Sign,saville,white,5060000000003,icon.svg,landscape
'''
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=product_template.csv'}
    )


@app.route('/api/templates/svg')
def download_svg_template():
    """Download a 100mm x 100mm SVG template for product graphics."""
    svg_content = '''<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" 
     width="100mm" height="100mm" 
     viewBox="0 0 100 100">
  <!-- 100mm x 100mm template for product icon/graphic -->
  <!-- Design your icon within this area -->
  
  <!-- Guide rectangle (remove before use) -->
  <rect x="0" y="0" width="100" height="100" 
        fill="none" stroke="#ccc" stroke-width="0.5" stroke-dasharray="2,2"/>
  
  <!-- Center crosshair (remove before use) -->
  <line x1="50" y1="0" x2="50" y2="100" stroke="#eee" stroke-width="0.25"/>
  <line x1="0" y1="50" x2="100" y2="50" stroke="#eee" stroke-width="0.25"/>
  
  <!-- Example icon placeholder - replace with your design -->
  <circle cx="50" cy="50" r="30" fill="none" stroke="#999" stroke-width="2"/>
  <text x="50" y="55" text-anchor="middle" font-family="Arial" font-size="8" fill="#999">
    Your Icon Here
  </text>
</svg>'''
    
    return Response(
        svg_content,
        mimetype='image/svg+xml',
        headers={'Content-Disposition': 'attachment; filename=icon_template_100mm.svg'}
    )


@app.route('/api/icons', methods=['GET'])
def list_icons():
    """List all available icon files."""
    icons_dir = Path(__file__).parent / "icons"
    if not icons_dir.exists():
        return jsonify([])
    
    icons = []
    for f in icons_dir.glob('*.svg'):
        icons.append({
            "filename": f.name,
            "path": str(f)
        })
    return jsonify(icons)


@app.route('/api/icons/<filename>', methods=['GET'])
def get_icon(filename):
    """Serve an icon file."""
    icons_dir = Path(__file__).parent / "icons"
    icon_path = icons_dir / filename
    
    if not icon_path.exists():
        return "Not found", 404
    
    return send_file(icon_path, mimetype='image/svg+xml')


@app.route('/api/icons/<filename>', methods=['DELETE'])
def delete_icon(filename):
    """Delete an icon file."""
    icons_dir = Path(__file__).parent / "icons"
    icon_path = icons_dir / filename
    
    if not icon_path.exists():
        return jsonify({"success": False, "error": "Icon not found"}), 404
    
    try:
        icon_path.unlink()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/icons/upload', methods=['POST'])
def upload_icon():
    """Upload an SVG icon file."""
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file provided"}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.svg'):
        return jsonify({"success": False, "error": "Only SVG files allowed"}), 400
    
    # Save to icons directory
    icons_dir = Path(__file__).parent / "icons"
    icons_dir.mkdir(exist_ok=True)
    
    # Use original filename or generate unique one
    filename = file.filename
    save_path = icons_dir / filename
    
    # If file exists, add number suffix
    counter = 1
    while save_path.exists():
        name, ext = filename.rsplit('.', 1)
        save_path = icons_dir / f"{name}_{counter}.{ext}"
        counter += 1
    
    file.save(save_path)
    
    return jsonify({
        "success": True,
        "filename": save_path.name,
        "path": str(save_path)
    })


# Simple in-memory cache for preview images
_preview_cache = {}

@app.route('/api/preview/clear-cache', methods=['POST'])
def clear_preview_cache():
    """Clear the preview image cache."""
    global _preview_cache
    _preview_cache = {}
    return jsonify({"success": True, "message": "Preview cache cleared"})

@app.route('/api/debug/icons')
def debug_icons():
    """Debug endpoint to check icons directory."""
    from pathlib import Path
    icons_dir = Path(__file__).parent / "icons"
    
    result = {
        "icons_dir": str(icons_dir),
        "exists": icons_dir.exists(),
        "is_dir": icons_dir.is_dir() if icons_dir.exists() else False,
        "files": []
    }
    
    if icons_dir.exists() and icons_dir.is_dir():
        try:
            result["files"] = [f.name for f in icons_dir.iterdir()]
        except Exception as e:
            result["error"] = str(e)
    
    return jsonify(result)

@app.route('/api/preview/<m_number>')
def preview_product(m_number):
    """Generate PNG preview for a product."""
    product = Product.get(m_number)
    if not product:
        return "Not found", 404
    
    # Check cache - key includes icon_files and updated_at to invalidate on changes
    cache_key = f"{m_number}_{product.get('icon_files', '')}_{product.get('updated_at', '')}"
    if cache_key in _preview_cache:
        return Response(_preview_cache[cache_key], mimetype='image/png')
    
    try:
        from image_generator import generate_product_image_preview
        import logging
        logging.info(f"Generating preview for {m_number} with icon: {product.get('icon_files', 'none')}")
        # Use low-res preview for thumbnails (scale=1 instead of scale=4)
        png_bytes = generate_product_image_preview(product)
        
        # Cache the result (limit cache size)
        if len(_preview_cache) > 100:
            # Remove oldest entries
            keys_to_remove = list(_preview_cache.keys())[:50]
            for k in keys_to_remove:
                del _preview_cache[k]
        _preview_cache[cache_key] = png_bytes
        
        return Response(png_bytes, mimetype='image/png')
    except Exception as e:
        import logging
        import traceback
        logging.error(f"Preview generation failed for {m_number}: {e}")
        logging.error(traceback.format_exc())
        # Fallback to placeholder SVG on error
        svg = f'''<svg xmlns="http://www.w3.org/2000/svg" width="300" height="200" viewBox="0 0 300 200">
            <rect width="300" height="200" fill="#f8d7da"/>
            <text x="150" y="90" text-anchor="middle" font-family="Arial" font-size="16" fill="#721c24">{m_number}</text>
            <text x="150" y="120" text-anchor="middle" font-family="Arial" font-size="12" fill="#721c24">Preview error</text>
            <text x="150" y="145" text-anchor="middle" font-family="Arial" font-size="10" fill="#999">{str(e)[:40]}</text>
        </svg>'''
        return Response(svg, mimetype='image/svg+xml')


@app.route('/api/analyze/products', methods=['POST'])
def analyze_products():
    """Analyze product images with AI to auto-populate theme and use cases."""
    import os
    import base64
    import logging
    from image_generator import generate_product_image
    
    data = request.json or {}
    sample_m_numbers = data.get('sample_m_numbers', [])
    
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return jsonify({"success": False, "error": "OPENAI_API_KEY not set"}), 400
    
    if not sample_m_numbers:
        return jsonify({"success": False, "error": "No sample images provided"}), 400
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        # Build message content with images
        content = []
        
        # Add sample images
        for m_number in sample_m_numbers[:5]:
            product = Product.get(m_number)
            if product:
                try:
                    png_bytes = generate_product_image(product, "main")
                    img_base64 = base64.b64encode(png_bytes).decode()
                    content.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{img_base64}"
                        }
                    })
                except Exception as e:
                    logging.warning(f"Failed to generate sample image for {m_number}: {e}")
        
        # Add analysis prompt
        content.append({
            "type": "text",
            "text": """Analyze these product images. They are self-adhesive aluminum signs.

Please provide:
1. THEME: A brief description of what type of sign this is (e.g., "No entry without permission sign", "Fire exit sign", "Warning sign"). Be specific about what the sign communicates.

2. USE_CASES: Where would this sign typically be used? List 2-4 locations or scenarios, separated by commas (e.g., "offices, warehouses, restricted areas, private property").

Respond in this exact format:
THEME: [your theme description]
USE_CASES: [comma-separated list of use cases]"""
        })
        
        # Call OpenAI GPT-4 Vision API
        response = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=500,
            messages=[
                {"role": "system", "content": "You are a product analyst. Analyze the sign images and identify what they communicate and where they would be used. Be concise and specific."},
                {"role": "user", "content": content}
            ]
        )
        
        result = response.choices[0].message.content
        
        # Parse the response
        theme = ""
        use_cases = ""
        
        for line in result.split('\n'):
            line = line.strip()
            if line.upper().startswith('THEME:'):
                theme = line[6:].strip()
            elif line.upper().startswith('USE_CASES:'):
                use_cases = line[10:].strip()
        
        return jsonify({
            "success": True,
            "theme": theme,
            "use_cases": use_cases,
            "raw_response": result
        })
        
    except Exception as e:
        logging.error(f"Failed to analyze products: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/generate/images', methods=['POST'])
def generate_images():
    """Generate product images for approved products."""
    from image_generator import generate_images_job
    
    products = Product.approved()
    if not products:
        products = Product.all()  # Fall back to all if none approved
    
    if not products:
        return jsonify({"error": "No products to generate"}), 400
    
    job_id = submit_job(
        f"Generate images for {len(products)} products",
        generate_images_job,
        products,
    )
    
    return jsonify({"success": True, "job_id": job_id, "count": len(products)})


@app.route('/api/generate/amazon-flatfile', methods=['POST'])
def generate_amazon_flatfile():
    """Generate Amazon flatfile XLSX in proper Amazon format."""
    import logging
    import os
    import re
    import json
    from datetime import datetime
    import openpyxl
    from openpyxl.utils import get_column_letter
    from config import PUBLIC_BASE_URL
    
    data = request.json or {}
    theme = data.get('theme', '')
    use_cases = data.get('use_cases', '')
    ai_content = data.get('ai_content', '')
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    # Load blank metadata from DB
    _blanks_cache = {b["slug"]: b for b in Blank.all()}
    def _dims(slug): b = _blanks_cache.get(slug, {}); return (b.get("width_mm", 100)/10, b.get("height_mm", 100)/10)
    def _code(slug): return _blanks_cache.get(slug, {}).get("amazon_code", "M")
    SIZE_PRICING = {
        "dracula": 10.99,
        "saville": 11.99,
        "dick": 12.99,
        "barzan": 15.99,
        "baby_jesus": 17.99,
    }
    COLOR_DISPLAY = {"silver": "Silver", "white": "White", "gold": "Gold"}
    
    try:
        # Parse AI content to extract bullet points and description
        # AI content format expected to have titles, bullets, descriptions
        default_bullets = [
            "Premium 1mm brushed aluminium construction with elegant finish provides professional appearance and long-lasting durability",
            "UV-resistant printing technology ensures text remains clear and legible for years, even under harsh sunlight exposure",
            "Self-adhesive backing allows quick peel and stick installation – NO drilling or tools required, ready to use immediately",
            "Fully weatherproof design withstands rain, snow, and temperature extremes, suitable for indoor and outdoor applications",
            "Clear, bold messaging ensures excellent visibility and compliance, perfect for maintaining security in restricted areas"
        ]
        default_description = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive. Professional quality signage for {use_cases}."
        default_search_terms = "sign warning notice metal plaque door wall sticker business office industrial safety weatherproof aluminium"
        
        # Derive parent SKU from theme
        parent_sku = theme.upper().replace(" ", "_").replace("-", "_")
        parent_sku = re.sub(r'[^A-Z0-9_]', '', parent_sku)
        if not parent_sku.endswith("_PARENT"):
            parent_sku = f"{parent_sku}_PARENT"
        
        # Amazon columns
        AMAZON_COLUMNS = [
            ("feed_product_type", "Product Type"),
            ("item_sku", "Seller SKU"),
            ("update_delete", "Update Delete"),
            ("brand_name", "Brand Name"),
            ("external_product_id", "Product ID"),
            ("external_product_id_type", "Product ID Type"),
            ("product_description", "Product Description"),
            ("part_number", "Manufacturer Part Number"),
            ("manufacturer", "Manufacturer"),
            ("item_name", "Item Name (aka Title)"),
            ("language_value", "Language"),
            ("recommended_browse_nodes", "Recommended Browse Nodes"),
            ("main_image_url", "Main Image URL"),
            ("other_image_url1", "Other Image Url1"),
            ("other_image_url2", "Other Image Url2"),
            ("other_image_url3", "Other Image Url3"),
            ("other_image_url4", "Other Image Url4"),
            ("other_image_url5", "Other Image Url5"),
            ("other_image_url6", "Other Image Url6"),
            ("other_image_url7", "Other Image Url7"),
            ("other_image_url8", "Other Image Url8"),
            ("relationship_type", "Relationship Type"),
            ("variation_theme", "Variation Theme"),
            ("parent_sku", "Parent SKU"),
            ("parent_child", "Parentage"),
            ("style_name", "Style Name"),
            ("bullet_point1", "Key Product Features"),
            ("bullet_point2", "Key Product Features"),
            ("bullet_point3", "Key Product Features"),
            ("bullet_point4", "Key Product Features"),
            ("bullet_point5", "Key Product Features"),
            ("generic_keywords", "Search Terms"),
            ("color_name", "Colour"),
            ("size_name", "Size"),
            ("color_map", "Colour Map"),
            ("size_map", "Size Map"),
            ("length_longer_edge", "Item Length Longer Edge"),
            ("length_longer_edge_unit_of_measure", "Item Length Unit"),
            ("width_shorter_edge", "Item Width Shorter Edge"),
            ("width_shorter_edge_unit_of_measure", "Item Width Unit"),
            ("batteries_required", "Batteries Required"),
            ("supplier_declared_dg_hz_regulation5", "Dangerous Goods"),
            ("country_of_origin", "Country/Region Of Origin"),
            ("list_price_with_tax", "List Price"),
            ("merchant_shipping_group_name", "Shipping Group"),
            ("condition_type", "Condition"),
            ("fulfillment_availability#1.fulfillment_channel_code", "Fulfillment Channel"),
            ("fulfillment_availability#1.quantity", "Quantity"),
            ("purchasable_offer[marketplace_id=A1F83G8C2ARO7P]#1.our_price#1.schedule#1.value_with_tax", "Price"),
        ]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Row 1: Template metadata
        ws.cell(row=1, column=1, value="TemplateType=fptcustom")
        ws.cell(row=1, column=2, value="Version=2025.1207")
        ws.cell(row=1, column=3, value="TemplateSignature=U0lHTkFHRQ==")
        
        # Row 2: Labels
        for col, (attr, label) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=2, column=col, value=label)
        
        # Row 3: Attribute names
        for col, (attr, label) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=3, column=col, value=attr)
        
        # Row 4: Parent row
        parent_title = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive"
        parent_data = {
            "feed_product_type": "signage",
            "item_sku": parent_sku,
            "update_delete": "Update",
            "brand_name": "NorthByNorthEast",
            "product_description": default_description,
            "part_number": parent_sku,
            "item_name": parent_title,
            "recommended_browse_nodes": "330215031",
            "variation_theme": "Size & Colour",
            "parent_child": "Parent",
            "generic_keywords": default_search_terms,
            "batteries_required": "No",
            "supplier_declared_dg_hz_regulation5": "Not Applicable",
            "country_of_origin": "Great Britain",
        }
        for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=4, column=col, value=parent_data.get(attr, ""))
        
        # Row 5+: Child products
        row_num = 5
        for product in all_products:
            m_number = product['m_number']
            size = product.get('size', 'saville').lower()
            color = product.get('color', 'silver').lower()
            ean = product.get('ean', '')
            
            dims = _dims(size)
            size_code = _code(size)
            price = SIZE_PRICING.get(size, 12.99)
            color_display = COLOR_DISPLAY.get(color, color.title())

            title = f"{theme} Sign – {dims[0]}x{dims[1]}cm Brushed Aluminium, Weatherproof, Self-Adhesive"
            style_name = f"{color_display}_{size_code}"
            main_image = f"/images/{m_number}/{m_number}-001.jpg"
            
            row_data = {
                "feed_product_type": "signage",
                "item_sku": m_number,
                "update_delete": "Update",
                "brand_name": "NorthByNorthEast",
                "external_product_id": ean,
                "external_product_id_type": "EAN" if ean else "",
                "product_description": default_description,
                "part_number": m_number,
                "manufacturer": "North By North East Print and Sign Limited",
                "item_name": title,
                "language_value": "en_GB",
                "recommended_browse_nodes": "330215031",
                "main_image_url": main_image,
                "other_image_url1": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-002.jpg",
                "other_image_url2": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-003.jpg",
                "other_image_url3": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-004.jpg",
                "other_image_url4": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-006.jpg",
                "relationship_type": "Variation",
                "variation_theme": "Size & Colour",
                "parent_sku": parent_sku,
                "parent_child": "Child",
                "style_name": style_name,
                "bullet_point1": default_bullets[0],
                "bullet_point2": default_bullets[1],
                "bullet_point3": default_bullets[2],
                "bullet_point4": default_bullets[3],
                "bullet_point5": default_bullets[4],
                "generic_keywords": default_search_terms,
                "color_name": color_display,
                "size_name": size_code,
                "color_map": color_display,
                "size_map": size_code,
                "length_longer_edge": str(dims[0]),
                "length_longer_edge_unit_of_measure": "Centimetres",
                "width_shorter_edge": str(dims[1]),
                "width_shorter_edge_unit_of_measure": "Centimetres",
                "batteries_required": "No",
                "supplier_declared_dg_hz_regulation5": "Not Applicable",
                "country_of_origin": "Great Britain",
                "list_price_with_tax": str(price),
                "merchant_shipping_group_name": "RM Tracked 48 Free, 24 -- £2.99, SD -- £7.99",
                "condition_type": "New",
                "fulfillment_availability#1.fulfillment_channel_code": "AMAZON_UK_RAFN",
                "fulfillment_availability#1.quantity": "5",
                "purchasable_offer[marketplace_id=A1F83G8C2ARO7P]#1.our_price#1.schedule#1.value_with_tax": str(price),
            }
            
            for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
                ws.cell(row=row_num, column=col, value=row_data.get(attr, ""))
            row_num += 1
        
        # Auto-adjust column widths
        for col in range(1, len(AMAZON_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(__file__).parent / f"amazon_flatfile_{timestamp}.xlsx"
        wb.save(output_path)
        
        logging.info(f"Amazon flatfile saved to {output_path} with {len(all_products)} products")
        
        return jsonify({
            "success": True,
            "product_count": len(all_products),
            "file_path": str(output_path),
            "message": f"Amazon flatfile saved: {output_path.name}"
        })
        
    except Exception as e:
        logging.error(f"Failed to generate Amazon flatfile: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/jobs')
def list_jobs():
    """List all background jobs."""
    jobs = get_all_jobs()
    return jsonify([job_to_dict(j) for j in jobs])


@app.route('/api/jobs/<job_id>')
def get_job_status(job_id):
    """Get status of a specific job."""
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job_to_dict(job))


@app.route('/api/generate/content', methods=['POST'])
def generate_content():
    """Generate AI content for products using OpenAI GPT-4 with sample images."""
    import os
    import base64
    import logging
    from image_generator import generate_product_image
    
    data = request.json or {}
    theme = data.get('theme', '')
    use_cases = data.get('use_cases', '')
    system_prompt = data.get('system_prompt', '')
    sample_m_numbers = data.get('sample_m_numbers', [])
    
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return jsonify({"success": False, "error": "OPENAI_API_KEY not set"}), 400
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        # Build product summary
        sizes = {}
        colors = set()
        descriptions = set()
        
        for p in all_products:
            size = p.get('size', 'unknown')
            color = p.get('color', 'unknown')
            sizes[size] = sizes.get(size, 0) + 1
            colors.add(color)
            if p.get('description'):
                descriptions.add(p['description'])
        
        size_dims = {
            'saville': '115x95mm (rectangular)',
            'dick': '140x90mm (rectangular)', 
            'barzan': '194x143mm (rectangular)',
            'dracula': '95mm diameter (circular)',
            'baby_jesus': '290x190mm (rectangular)'
        }
        
        summary = f"Total Products: {len(all_products)}\n"
        summary += f"Colors: {', '.join(colors)}\n"
        summary += "Sizes:\n"
        for size, count in sizes.items():
            dims = size_dims.get(size, 'unknown dimensions')
            summary += f"  - {size}: {count} products - {dims}\n"
        summary += f"Product Types: {len(descriptions)} unique designs"
        
        # Build message content with images for GPT-4 Vision
        content = []
        
        # Add sample images (use cached previews if available, otherwise generate)
        for m_number in sample_m_numbers[:3]:  # Reduced from 5 to 3 for faster processing
            product = Product.get(m_number)
            if product:
                try:
                    # Try to use cached preview first (scale=1), fall back to full generation
                    cache_key = f"{m_number}_{product.get('icon_files', '')}_{product.get('updated_at', '')}"
                    if cache_key in _preview_cache:
                        png_bytes = _preview_cache[cache_key]
                        logging.info(f"Using cached preview for {m_number}")
                    else:
                        from image_generator import generate_product_image_preview
                        png_bytes = generate_product_image_preview(product)
                        logging.info(f"Generated preview for {m_number}")
                    
                    img_base64 = base64.b64encode(png_bytes).decode()
                    content.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{img_base64}"
                        }
                    })
                except Exception as e:
                    logging.warning(f"Failed to generate sample image for {m_number}: {e}")
        
        # Add text prompt
        user_prompt = f"""=== PRODUCT SUMMARY ===
{summary}

=== USER INPUT ===
Theme: {theme or '(not specified)'}
Use Cases: {use_cases or '(not specified)'}

Please generate Amazon marketplace content for these products. Include:
1. Product titles (under 200 characters)
2. 5 bullet points per product
3. Product descriptions

Remember: These products come in MULTIPLE sizes and shapes as shown in the images and summary above."""

        content.append({"type": "text", "text": user_prompt})
        
        # Call OpenAI GPT-4 Vision API
        response = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=4096,
            messages=[
                {"role": "system", "content": system_prompt or "You are an expert product content writer for Amazon marketplace listings."},
                {"role": "user", "content": content}
            ]
        )
        
        generated_content = response.choices[0].message.content
        
        return jsonify({
            "success": True,
            "content": generated_content
        })
        
    except Exception as e:
        logging.error(f"Failed to generate content: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/generate/full', methods=['POST'])
def generate_full():
    """Run full pipeline."""
    def stream():
        yield "Full pipeline not yet implemented in web version.\n"
    return Response(stream(), mimetype='text/plain')


@app.route('/api/ebay/publish', methods=['POST'])
def publish_to_ebay():
    """Publish approved products to eBay via API with auto-promotion."""
    import logging
    
    data = request.json or {}
    promote = data.get('with_ads', True)
    ad_rate = data.get('ad_rate', '5.0')
    dry_run = data.get('dry_run', False)
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    if not products:
        return jsonify({"success": False, "error": "No products to publish"}), 400
    
    try:
        from ebay_api import create_ebay_listing, load_policy_ids
        policy_ids = load_policy_ids()
    except FileNotFoundError as e:
        logging.warning(f"eBay policies not configured: {e}")
        return jsonify({"success": False, "error": "eBay policies not configured. Run ebay_setup_policies.py first."}), 400
    except ImportError as e:
        logging.warning(f"eBay API module not available: {e}")
        return jsonify({"success": False, "error": "eBay API module not available"}), 400
    except Exception as e:
        logging.error(f"Error loading eBay policies: {e}")
        return jsonify({"success": False, "error": f"eBay setup error: {str(e)}"}), 400
    
    try:
        logging.info(f"Publishing {len(products)} products to eBay (promote={promote})")
        listing_id = create_ebay_listing(
            products=products,
            policy_ids=policy_ids,
            promote=promote,
            ad_rate=ad_rate,
            dry_run=dry_run,
        )
        
        if listing_id:
            result = {
                "success": True,
                "listing_id": listing_id,
                "count": len(products),
                "promoted": promote and listing_id != "DRY_RUN",
            }
            if listing_id != "DRY_RUN":
                result["url"] = f"https://www.ebay.co.uk/itm/{listing_id}"
            logging.info(f"eBay listing created: {listing_id}")
            return jsonify(result)
        else:
            return jsonify({"success": False, "error": "Failed to create listing - no listing ID returned"}), 500
    except Exception as e:
        logging.error(f"eBay publish error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/flatfile-preview')
def flatfile_preview():
    """Return Amazon flatfile data as JSON for preview table."""
    import os
    from config import PUBLIC_BASE_URL
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    _blanks_cache = {b["slug"]: b for b in Blank.all()}
    def _dims(slug): b = _blanks_cache.get(slug, {}); return (b.get("width_mm", 100)/10, b.get("height_mm", 100)/10)
    def _code(slug): return _blanks_cache.get(slug, {}).get("amazon_code", "M")
    SIZE_PRICING = {
        "dracula": 10.99, "saville": 11.99, "dick": 12.99,
        "barzan": 15.99, "baby_jesus": 17.99,
    }
    COLOR_DISPLAY = {"silver": "Silver", "white": "White", "gold": "Gold"}

    # Get theme from first product
    theme = all_products[0].get('description', 'Sign') if all_products else 'Sign'
    parent_sku = theme.upper().replace(" ", "_").replace("-", "_")
    import re
    parent_sku = re.sub(r'[^A-Z0-9_]', '', parent_sku)
    if not parent_sku.endswith("_PARENT"):
        parent_sku = f"{parent_sku}_PARENT"
    
    # Headers for display - ALL columns from flatfile (including lifestyle image)
    headers = ['item_sku', 'parent_child', 'item_name', 'color_name', 'size_name', 'external_product_id', 
               'list_price', 'main_image_url', 'other_image_url1', 'other_image_url2', 'other_image_url3', 
               'other_image_url4', 'other_image_url5', 'bullet_point1', 'bullet_point2', 'generic_keywords']
    
    rows = []
    
    default_bullets = [
        "Premium 1mm brushed aluminium construction",
        "UV-resistant printing technology",
    ]
    
    # Parent row
    parent_title = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive"
    rows.append({
        'item_sku': parent_sku,
        'parent_child': 'Parent',
        'item_name': parent_title,
        'color_name': '',
        'size_name': '',
        'external_product_id': '',
        'list_price': '',
        'main_image_url': '',
        'other_image_url1': '',
        'other_image_url2': '',
        'other_image_url3': '',
        'other_image_url4': '',
        'bullet_point1': '',
        'bullet_point2': '',
        'generic_keywords': ''
    })
    
    # Child rows
    for product in all_products:
        m_number = product['m_number']
        size = product.get('size', 'saville').lower()
        color = product.get('color', 'silver').lower()
        ean = product.get('ean', '')
        
        dims = _dims(size)
        size_code = _code(size)
        price = SIZE_PRICING.get(size, 12.99)
        color_display = COLOR_DISPLAY.get(color, color.title())

        title = f"{theme} Sign – {dims[0]}x{dims[1]}cm Brushed Aluminium"
        main_image = f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-001.jpg"
        
        rows.append({
            'item_sku': m_number,
            'parent_child': 'Child',
            'item_name': title,
            'color_name': color_display,
            'size_name': size_code,
            'external_product_id': str(ean) if ean else '',
            'list_price': f"£{price:.2f}",
            'main_image_url': f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-001.jpg",
            'other_image_url1': f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-002.jpg",
            'other_image_url2': f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-003.jpg",
            'other_image_url3': f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-004.jpg",
            'other_image_url4': f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-006.jpg",
            'bullet_point1': default_bullets[0],
            'bullet_point2': default_bullets[1],
            'generic_keywords': 'sign warning notice metal plaque weatherproof'
        })
    
    return jsonify({
        "success": True,
        "headers": headers,
        "rows": rows
    })


@app.route('/api/export/amazon-flatfile-download', methods=['POST'])
def download_amazon_flatfile():
    """Generate and download Amazon flatfile XLSX."""
    import os
    import re
    from io import BytesIO
    from datetime import datetime
    import openpyxl
    from openpyxl.utils import get_column_letter
    from config import PUBLIC_BASE_URL
    
    data = request.json or {}
    theme = data.get('theme', '')
    use_cases = data.get('use_cases', '')
    
    all_products = Product.all()
    if not all_products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
    _blanks_cache = {b["slug"]: b for b in Blank.all()}
    def _dims(slug): b = _blanks_cache.get(slug, {}); return (b.get("width_mm", 100)/10, b.get("height_mm", 100)/10)
    def _code(slug): return _blanks_cache.get(slug, {}).get("amazon_code", "M")
    SIZE_PRICING = {"dracula": 10.99, "saville": 11.99, "dick": 12.99, "barzan": 15.99, "baby_jesus": 17.99}
    COLOR_DISPLAY = {"silver": "Silver", "white": "White", "gold": "Gold"}

    if not theme:
        theme = all_products[0].get('description', 'Sign')
    
    parent_sku = re.sub(r'[^A-Z0-9_]', '', theme.upper().replace(" ", "_").replace("-", "_"))
    if not parent_sku.endswith("_PARENT"):
        parent_sku = f"{parent_sku}_PARENT"
    
    default_bullets = [
        "Premium 1mm brushed aluminium construction with elegant finish",
        "UV-resistant printing technology ensures text remains clear and legible",
        "Self-adhesive backing allows quick peel and stick installation",
        "Fully weatherproof design withstands rain, snow, and temperature extremes",
        "Clear, bold messaging ensures excellent visibility and compliance"
    ]
    default_description = f"{theme} Sign – Brushed Aluminium, Weatherproof, Self-Adhesive."
    default_search_terms = "sign warning notice metal plaque weatherproof aluminium"
    
    AMAZON_COLUMNS = [
        ("feed_product_type", "Product Type"), ("item_sku", "Seller SKU"), ("update_delete", "Update Delete"),
        ("brand_name", "Brand Name"), ("external_product_id", "Product ID"), ("external_product_id_type", "Product ID Type"),
        ("product_description", "Product Description"), ("part_number", "Part Number"), ("manufacturer", "Manufacturer"),
        ("item_name", "Item Name"), ("recommended_browse_nodes", "Browse Nodes"),
        ("main_image_url", "Main Image URL"), ("other_image_url1", "Other Image 1"), ("other_image_url2", "Other Image 2"),
        ("other_image_url3", "Other Image 3"), ("other_image_url4", "Other Image 4"),
        ("relationship_type", "Relationship Type"), ("variation_theme", "Variation Theme"),
        ("parent_sku", "Parent SKU"), ("parent_child", "Parentage"), ("style_name", "Style Name"),
        ("bullet_point1", "Bullet 1"), ("bullet_point2", "Bullet 2"), ("bullet_point3", "Bullet 3"),
        ("bullet_point4", "Bullet 4"), ("bullet_point5", "Bullet 5"), ("generic_keywords", "Search Terms"),
        ("color_name", "Colour"), ("size_name", "Size"), ("color_map", "Colour Map"), ("size_map", "Size Map"),
        ("list_price_with_tax", "List Price"), ("country_of_origin", "Country"),
    ]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Headers
    for col, (attr, label) in enumerate(AMAZON_COLUMNS, 1):
        ws.cell(row=1, column=col, value=label)
        ws.cell(row=2, column=col, value=attr)
    
    # Parent row
    parent_data = {"feed_product_type": "signage", "item_sku": parent_sku, "update_delete": "Update",
        "brand_name": "NorthByNorthEast", "item_name": f"{theme} Sign – Brushed Aluminium",
        "variation_theme": "Size & Colour", "parent_child": "Parent", "country_of_origin": "Great Britain"}
    for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
        ws.cell(row=3, column=col, value=parent_data.get(attr, ""))
    
    # Child rows
    row_num = 4
    for product in all_products:
        m_number = product['m_number']
        size = product.get('size', 'saville').lower()
        color = product.get('color', 'silver').lower()
        ean = product.get('ean', '')
        dims = _dims(size)
        size_code = _code(size)
        price = SIZE_PRICING.get(size, 12.99)
        color_display = COLOR_DISPLAY.get(color, color.title())
        
        row_data = {
            "feed_product_type": "signage", "item_sku": m_number, "update_delete": "Update",
            "brand_name": "NorthByNorthEast", "external_product_id": ean,
            "external_product_id_type": "EAN" if ean else "", "product_description": default_description,
            "part_number": m_number, "manufacturer": "North By North East Print and Sign Limited",
            "item_name": f"{theme} Sign – {dims[0]}x{dims[1]}cm Brushed Aluminium",
            "recommended_browse_nodes": "330215031",
            "main_image_url": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-001.jpg",
            "other_image_url1": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-002.jpg",
            "other_image_url2": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-003.jpg",
            "other_image_url3": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-004.jpg",
            "other_image_url4": f"{PUBLIC_BASE_URL}/images/{m_number}/{m_number}-006.jpg",
            "relationship_type": "Variation", "variation_theme": "Size & Colour",
            "parent_sku": parent_sku, "parent_child": "Child", "style_name": f"{color_display}_{size_code}",
            "bullet_point1": default_bullets[0], "bullet_point2": default_bullets[1],
            "bullet_point3": default_bullets[2], "bullet_point4": default_bullets[3],
            "bullet_point5": default_bullets[4], "generic_keywords": default_search_terms,
            "color_name": color_display, "size_name": size_code, "color_map": color_display,
            "size_map": size_code, "list_price_with_tax": str(price), "country_of_origin": "Great Britain",
        }
        for col, (attr, _) in enumerate(AMAZON_COLUMNS, 1):
            ws.cell(row=row_num, column=col, value=row_data.get(attr, ""))
        row_num += 1
    
    for col in range(1, len(AMAZON_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'amazon_flatfile_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')


@app.route('/api/export/etsy-download', methods=['POST'])
def download_etsy_file():
    """Generate and download Etsy shop uploader XLSX."""
    from io import BytesIO
    from datetime import datetime
    
    try:
        from export_etsy import generate_etsy_xlsx
        from config import PUBLIC_BASE_URL
        
        products = Product.approved()
        if not products:
            products = Product.all()
        
        if not products:
            return jsonify({"success": False, "error": "No products found"}), 400
        
        xlsx_bytes = generate_etsy_xlsx(products, PUBLIC_BASE_URL)
        
        return send_file(BytesIO(xlsx_bytes), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'etsy_shop_uploader_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/amazon', methods=['POST'])
@app.route('/api/export/flatfile', methods=['POST'])
def export_flatfile():
    """Export Amazon flatfile for approved products."""
    from io import BytesIO
    import openpyxl
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Headers
    headers = ['m_number', 'description', 'size', 'color', 'ean', 'qa_status']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, p in enumerate(products, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=p.get(header, ''))
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'amazon_flatfile_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )


@app.route('/api/export/ebay', methods=['POST'])
def export_ebay():
    """Export eBay File Exchange CSV."""
    from export_ebay import generate_ebay_csv
    from config import PUBLIC_BASE_URL
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    csv_content = generate_ebay_csv(products, PUBLIC_BASE_URL)
    
    return Response(
        csv_content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=ebay_listings_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'}
    )


@app.route('/api/export/etsy', methods=['POST'])
def export_etsy():
    """Export Etsy Shop Uploader XLSX."""
    import logging
    from datetime import datetime
    
    try:
        from export_etsy import generate_etsy_xlsx
        from config import PUBLIC_BASE_URL
        
        products = Product.approved()
        if not products:
            products = Product.all()
        
        if not products:
            return jsonify({"success": False, "error": "No products found"}), 400
        
        xlsx_bytes = generate_etsy_xlsx(products, PUBLIC_BASE_URL)
        
        # Save to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = Path(__file__).parent / f"etsy_shop_uploader_{timestamp}.xlsx"
        with open(output_path, 'wb') as f:
            f.write(xlsx_bytes)
        
        return jsonify({
            "success": True,
            "product_count": len(products),
            "file_path": str(output_path),
            "message": f"Etsy file saved: {output_path.name}"
        })
    except Exception as e:
        logging.error(f"Failed to generate Etsy file: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/lifestyle-background', methods=['POST'])
def generate_lifestyle_background():
    """Generate a lifestyle background image using DALL-E."""
    import os
    import logging
    import traceback
    
    try:
        import requests
        from datetime import datetime
        
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            logging.error("OPENAI_API_KEY not set in environment")
            return jsonify({"success": False, "error": "OPENAI_API_KEY not set. Please set the environment variable."}), 400
        
        # Get theme from first product or request
        data = request.json or {}
        theme = data.get('theme', '')
        
        if not theme:
            products = Product.all()
            if products:
                theme = products[0].get('description', 'safety sign')
        
        logging.info(f"Generating lifestyle background for theme: {theme}")
        
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        prompt = """A minimalist photograph of a plain concrete or painted wall in a modern building.
The wall is completely blank and empty with nothing on it at all. 
Simple industrial or office interior with soft natural lighting from a window.
Focus on texture and clean lines. No objects, no decorations, no artwork, no fixtures on the wall.
Architectural photography style, shallow depth of field, professional quality."""
        
        logging.info("Calling DALL-E 3 API...")
        response = client.images.generate(
            model="dall-e-3",
            prompt=prompt,
            size="1024x1024",
            quality="standard",
            n=1,
        )
        
        image_url = response.data[0].url
        logging.info(f"DALL-E returned image URL: {image_url[:50]}...")
        
        # Download and save the image locally
        img_response = requests.get(image_url, timeout=60)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        bg_path = Path(__file__).parent / f"lifestyle_background_{timestamp}.png"
        with open(bg_path, 'wb') as f:
            f.write(img_response.content)
        
        logging.info(f"Lifestyle background saved to {bg_path}")
        
        return jsonify({
            "success": True,
            "background_url": f"/api/export/lifestyle-background/preview?file={bg_path.name}",
            "file_path": str(bg_path)
        })
        
    except ImportError as e:
        logging.error(f"Import error: {e}")
        return jsonify({"success": False, "error": f"Missing dependency: {e}"}), 500
    except Exception as e:
        logging.error(f"Failed to generate lifestyle background: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/lifestyle-background/preview')
def preview_lifestyle_background():
    """Serve the lifestyle background image."""
    file_name = request.args.get('file')
    if not file_name:
        return "File not specified", 400
    
    file_path = Path(__file__).parent / file_name
    if not file_path.exists():
        return "File not found", 404
    
    return send_file(file_path, mimetype='image/png')


@app.route('/api/export/lifestyle-images', methods=['POST'])
def generate_lifestyle_images():
    """Generate lifestyle images by overlaying product PNGs on the background and upload to R2."""
    import logging
    import os
    from PIL import Image
    from io import BytesIO
    from image_generator import generate_product_image
    
    data = request.json or {}
    background_url = data.get('background_url', '')
    
    logging.info(f"Lifestyle images request - background_url: {background_url}")
    
    if not background_url:
        return jsonify({"success": False, "error": "No background URL provided"}), 400
    
    # Extract file name from URL - handle both /api/export/file?file=X and direct paths
    if 'file=' in background_url:
        file_name = background_url.split('file=')[-1]
    elif background_url.startswith('/'):
        # Direct path like /api/export/file?file=lifestyle_background_xxx.png
        file_name = background_url.split('/')[-1]
    else:
        file_name = background_url
    
    logging.info(f"Extracted file_name: {file_name}")
    bg_path = Path(__file__).parent / file_name
    logging.info(f"Looking for background at: {bg_path}")
    
    if not bg_path.exists():
        # Try to find the most recent lifestyle background
        bg_files = list(Path(__file__).parent.glob("lifestyle_background_*.png"))
        if bg_files:
            bg_path = max(bg_files, key=lambda p: p.stat().st_mtime)
            logging.info(f"Using most recent background: {bg_path}")
        else:
            logging.error(f"Background file not found: {bg_path}")
            return jsonify({"success": False, "error": f"Background file not found: {file_name}"}), 400
    
    products = Product.all()
    if not products:
        return jsonify({"success": False, "error": "No products found"}), 400
    
        
    try:
        # Load background image
        background = Image.open(bg_path).convert('RGBA')
        bg_width, bg_height = background.size
        
        count = 0
        images_data = []
        
        from local_storage import save_image as save_to_disk

        for product in products:
            try:
                m_number = product['m_number']
                logging.info(f"Creating lifestyle image for {m_number}...")

                from image_generator import generate_transparent_product_image
                png_bytes = generate_transparent_product_image(product)
                product_img = Image.open(BytesIO(png_bytes)).convert('RGBA')

                target_width = int(bg_width * 0.4)
                ratio = target_width / product_img.width
                target_height = int(product_img.height * ratio)
                product_img = product_img.resize((target_width, target_height), Image.Resampling.LANCZOS)

                composite = background.copy()
                x_pos = int(bg_width * 0.55 - target_width // 2)
                y_pos = int(bg_height * 0.45 - target_height // 2)
                composite.paste(product_img, (x_pos, y_pos), product_img)

                img_bytes = BytesIO()
                composite.convert('RGB').save(img_bytes, 'JPEG', quality=90)
                save_to_disk(img_bytes.getvalue(), f"{m_number}/{m_number}_lifestyle.jpg", "image/jpeg")

                images_data.append({
                    'm_number': m_number,
                    'url': f"/api/export/lifestyle-preview/{m_number}",
                })
                count += 1
                
            except Exception as e:
                import traceback
                logging.error(f"Failed to create lifestyle for {product['m_number']}: {e}")
                logging.error(traceback.format_exc())
        
        return jsonify({
            "success": True,
            "count": count,
            "images": images_data,
            "message": f"Created {count} lifestyle images"
        })
        
    except Exception as e:
        logging.error(f"Failed to generate lifestyle images: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/lifestyle-preview/<m_number>')
def preview_lifestyle_image(m_number):
    """Serve a lifestyle image preview from local storage."""
    for ext in ("jpg", "png"):
        file_path = IMAGES_DIR / m_number / f"{m_number}_lifestyle.{ext}"
        if file_path.exists():
            return send_file(file_path)
    return "File not found", 404


@app.route('/api/open-folder', methods=['POST'])
def open_folder():
    """Open a folder in the system file explorer."""
    import subprocess
    import os
    
    data = request.json or {}
    folder_type = data.get('type', 'm_number')
    
    # Define folder paths
    FOLDERS = {
        'm_number': r"G:\My Drive\001 NBNE\001 M",
        'exports': r"G:\My Drive\003 APPS\019 - AMAZON PUBLISHER REV 2.0\exports",
    }
    
    folder_path = FOLDERS.get(folder_type)
    if not folder_path:
        return jsonify({"success": False, "error": f"Unknown folder type: {folder_type}"}), 400
    
    if not os.path.exists(folder_path):
        # Try to create it
        try:
            os.makedirs(folder_path, exist_ok=True)
        except Exception as e:
            return jsonify({"success": False, "error": f"Folder does not exist and could not be created: {folder_path}"}), 400
    
    try:
        # Open folder in Windows Explorer
        subprocess.Popen(['explorer', folder_path])
        return jsonify({"success": True, "path": folder_path})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/export/m-folders', methods=['POST'])
def export_m_folders_json():
    """Generate M Number folders ZIP and stream it directly to the browser."""
    import logging
    from datetime import datetime
    from io import BytesIO
    from export_images import generate_m_number_folder_zip

    products = Product.approved()
    if not products:
        products = Product.all()

    if not products:
        return jsonify({"success": False, "error": "No products found"}), 400

    try:
        zip_bytes = generate_m_number_folder_zip(products)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return send_file(
            BytesIO(zip_bytes),
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"m_number_folders_{timestamp}.zip"
        )
    except Exception as e:
        logging.error(f"Failed to generate M folders ZIP: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/upload-images-to-r2', methods=['POST'])
@app.route('/api/generate/save-images', methods=['POST'])
def save_product_images():
    """Generate all product images and save to local storage."""
    import logging
    import traceback
    from io import BytesIO
    from PIL import Image
    from image_generator import generate_product_image

    products = Product.all()
    if not products:
        return jsonify({"success": False, "error": "No products found"}), 400

    IMAGE_TYPES = [
        ("main", "001"),
        ("dimensions", "002"),
        ("peel_and_stick", "003"),
        ("rear", "004"),
    ]

    from local_storage import save_image as save_to_disk
    from concurrent.futures import ThreadPoolExecutor, as_completed
    total_saved = 0
    errors = []

    def _render_one(product, img_type, img_num):
        m_number = product["m_number"]
        png_bytes = generate_product_image(product, img_type)
        img = Image.open(BytesIO(png_bytes))
        if img.mode == "RGBA":
            bg = Image.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        else:
            img = img.convert("RGB")
        MAX_DIM = 2000
        if img.width > MAX_DIM or img.height > MAX_DIM:
            ratio = min(MAX_DIM / img.width, MAX_DIM / img.height)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
        buf = BytesIO()
        img.save(buf, "JPEG", quality=85)
        save_to_disk(buf.getvalue(), f"{m_number}/{m_number}-{img_num}.jpg", "image/jpeg")
        return m_number, img_type

    tasks = [
        (product, img_type, img_num)
        for product in products
        for img_type, img_num in IMAGE_TYPES
    ]

    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {pool.submit(_render_one, p, t, n): (p["m_number"], t) for p, t, n in tasks}
        for fut in as_completed(futures):
            m_number, img_type = futures[fut]
            try:
                fut.result()
                total_saved += 1
            except Exception as e:
                errors.append(f"{m_number} {img_type}: {e}")
                logging.error(traceback.format_exc())

    return jsonify({
        "success": total_saved > 0,
        "total_saved": total_saved,
        "products": len(products),
        "errors": errors[:20] if errors else [],
        "message": f"Saved {total_saved} images for {len(products)} products." + (f" ({len(errors)} errors)" if errors else ""),
    })


@app.route('/api/export/images/<m_number>', methods=['GET'])
def export_product_images(m_number):
    """Get product image(s). If type param provided, return single PNG. Otherwise return ZIP of all."""
    from io import BytesIO
    from image_generator import generate_product_image
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    # If type parameter provided, return single image
    img_type = request.args.get('type')
    if img_type:
        valid_types = ['main', 'dimensions', 'peel_and_stick', 'rear']
        if img_type not in valid_types:
            return jsonify({"error": f"Invalid type. Must be one of: {valid_types}"}), 400
        
        try:
            png_bytes = generate_product_image(product, img_type)
            return send_file(
                BytesIO(png_bytes),
                mimetype='image/png',
                download_name=f'{m_number}_{img_type}.png'
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    # No type param - return ZIP of all images
    from export_images import generate_single_product_zip
    zip_bytes = generate_single_product_zip(product)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'{m_number}_images.zip'
    )


@app.route('/api/export/images', methods=['POST'])
def export_all_images():
    """Download all product images as ZIP (approved products)."""
    from export_images import generate_images_zip
    from io import BytesIO
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    if not products:
        return jsonify({"error": "No products to export"}), 400
    
    zip_bytes = generate_images_zip(products)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'product_images_{datetime.now().strftime("%Y%m%d_%H%M")}.zip'
    )


@app.route('/api/export/m-number-folders/<m_number>', methods=['GET'])
def export_m_number_folder(m_number):
    """Download M Number folder with full structure for staff (single product)."""
    from export_images import generate_single_m_number_folder_zip
    from io import BytesIO
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    zip_bytes = generate_single_m_number_folder_zip(product)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'{m_number}_folder.zip'
    )


@app.route('/api/export/m-number-folders', methods=['POST'])
def export_all_m_number_folders():
    """Download all M Number folders with full structure for staff."""
    from export_images import generate_m_number_folder_zip
    from io import BytesIO
    
    products = Product.approved()
    if not products:
        products = Product.all()
    
    if not products:
        return jsonify({"error": "No products to export"}), 400
    
    zip_bytes = generate_m_number_folder_zip(products)
    
    return send_file(
        BytesIO(zip_bytes),
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'm_number_folders_{datetime.now().strftime("%Y%m%d_%H%M")}.zip'
    )


@app.route('/api/products/<m_number>/scale', methods=['PATCH'])
def update_product_scale(m_number):
    """Update icon_scale and text_scale for a product (QA tuning)."""
    data = request.json
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    updates = {}
    if 'icon_scale' in data:
        updates['icon_scale'] = float(data['icon_scale'])
    if 'text_scale' in data:
        updates['text_scale'] = float(data['text_scale'])
    
    if updates:
        Product.update(m_number, updates)
    
    return jsonify({"success": True, "updates": updates})


@app.route('/api/products/<m_number>/position', methods=['PATCH'])
def update_product_position(m_number):
    """Update icon_offset_x and icon_offset_y for a product (QA positioning)."""
    data = request.json
    
    product = Product.get(m_number)
    if not product:
        return jsonify({"error": "Product not found"}), 404
    
    updates = {}
    if 'icon_offset_x' in data:
        updates['icon_offset_x'] = float(data['icon_offset_x'])
    if 'icon_offset_y' in data:
        updates['icon_offset_y'] = float(data['icon_offset_y'])
    
    if updates:
        Product.update(m_number, updates)
    
    return jsonify({"success": True, "updates": updates})


# ── Image serving ──────────────────────────────────────────────────────────────

@app.route('/images/<path:filename>')
def serve_image(filename):
    """Serve product images from local storage."""
    # Prevent path traversal
    safe = Path(IMAGES_DIR / filename).resolve()
    if not str(safe).startswith(str(IMAGES_DIR.resolve())):
        return "Access denied", 403
    return send_from_directory(IMAGES_DIR, filename)


# ── Blanks API ─────────────────────────────────────────────────────────────────

@app.route('/api/blanks', methods=['GET'])
def list_blanks():
    """List all blanks."""
    active_only = request.args.get('active') == '1'
    return jsonify(Blank.all(active_only=active_only))


@app.route('/api/blanks/<slug>', methods=['GET'])
def get_blank(slug):
    blank = Blank.get(slug)
    if not blank:
        return jsonify({"error": "Blank not found"}), 404
    return jsonify(blank)


@app.route('/api/blanks', methods=['POST'])
def create_blank():
    """Register a new blank."""
    data = request.json
    required = {"slug", "display", "width_mm", "height_mm", "sign_x", "sign_y", "sign_w", "sign_h"}
    missing = required - set(data.keys())
    if missing:
        return jsonify({"error": f"Missing fields: {missing}"}), 400
    if Blank.get(data["slug"]):
        return jsonify({"error": f"Blank '{data['slug']}' already exists"}), 409
    Blank.create(data)
    return jsonify({"success": True, "slug": data["slug"]}), 201


@app.route('/api/blanks/<slug>', methods=['PATCH'])
def update_blank(slug):
    """Update a blank's configuration."""
    if not Blank.get(slug):
        return jsonify({"error": "Blank not found"}), 404
    Blank.update(slug, request.json)
    return jsonify({"success": True})


# ── Sales feedback loop ────────────────────────────────────────────────────────

def _parse_sales_csv(file_content: str) -> list[dict]:
    """Parse an Amazon Business Report CSV into a list of row dicts."""
    import csv, io

    def _n(v):
        v = str(v).replace(',', '').replace('£', '').replace('%', '').strip()
        try:
            return float(v)
        except ValueError:
            return 0.0

    rows = []
    reader = csv.DictReader(io.StringIO(file_content))
    for row in reader:
        asin = row.get('(Child) ASIN', '').strip()
        sku  = row.get('SKU', '').strip()
        if not asin and not sku:
            continue
        rows.append({
            'asin':        asin,
            'parent_asin': row.get('(Parent) ASIN', '').strip(),
            'sku':         sku,
            'title':       row.get('Title', '').strip(),
            'sessions':    _n(row.get('Sessions \u2013 Total', 0)),
            'units':       _n(row.get('Units ordered', 0)),
            'revenue':     _n(row.get('Ordered Product Sales', 0)),
            'cvr':         _n(row.get('Unit Session Percentage', 0)),
            'buy_box_pct': _n(row.get('Featured Offer (Buy Box) percentage', 0)),
        })
    return rows


def _infer_category(title: str) -> str:
    t = title.lower()
    if any(w in t for w in ('push', 'pull', 'door sign')):   return 'Push/Pull Door Signs'
    if any(w in t for w in ('smok', 'vap')):                  return 'No Smoking/Vaping'
    if 'park' in t:                                           return 'Parking'
    if any(w in t for w in ('bereavement', 'memorial', 'remembrance', 'tribute', "in memory")): return 'Memorial'
    if any(w in t for w in ('dog', 'pet', 'animal')):         return 'Dogs/Pets'
    if any(w in t for w in ('staff only', 'no entry', 'restricted', 'authorised', 'authorized')): return 'Access/Restricted'
    if any(w in t for w in ('private', 'property', 'trespass')): return 'Private Property'
    if any(w in t for w in ('caller', 'canvass', 'sales people', 'cold call')): return 'No Cold Callers'
    if any(w in t for w in ('cctv', 'surveillance', 'camera', 'security')):    return 'CCTV/Security'
    if any(w in t for w in ('fire', 'emergency', 'evacuation')):               return 'Fire Safety'
    if any(w in t for w in ('photog', 'filming', 'recording')):                return 'Photography/Filming'
    if any(w in t for w in ('toilet', 'bathroom', 'wc', 'flush')):             return 'Bathroom/WC'
    if any(w in t for w in ('caution', 'warning', 'danger', 'hazard')):        return 'Hazard/Warning'
    if 'parcel' in t or 'delivery' in t:                                        return 'Delivery/Parcel'
    return 'Other'


@app.route('/api/sales/import', methods=['POST'])
def import_sales():
    """Accept an Amazon Business Report CSV upload and store the sales data."""
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files['file']
    report_start = request.form.get('report_start', '').strip()
    report_end   = request.form.get('report_end', '').strip()

    if not report_start or not report_end:
        return jsonify({"error": "report_start and report_end are required (YYYY-MM-DD)"}), 400

    if SalesImport.import_exists(report_start, report_end):
        return jsonify({"error": f"Report for {report_start}–{report_end} already imported"}), 409

    content = f.read().decode('utf-8-sig')
    rows = _parse_sales_csv(content)
    if not rows:
        return jsonify({"error": "No data rows found in CSV"}), 400

    import_id = SalesImport.create(
        filename=f.filename,
        report_start=report_start,
        report_end=report_end,
        row_count=len(rows),
        imported_by=session.get('user_email', 'unknown'),
    )

    for row in rows:
        row['import_id']    = import_id
        row['report_start'] = report_start
        row['report_end']   = report_end

    SalesData.bulk_insert(rows)
    return jsonify({"ok": True, "import_id": import_id, "rows": len(rows)})


@app.route('/api/sales/performance')
def sales_performance():
    """Return aggregated performance data across all imports."""
    from collections import defaultdict

    top = SalesData.top_performers(limit=100)
    all_rows = SalesData.category_summary()
    imports = SalesImport.list_all()

    # Category rollup
    cats = defaultdict(lambda: {'units': 0, 'revenue': 0, 'products': 0, 'sessions': 0})
    for r in all_rows:
        cat = _infer_category(r['title'])
        cats[cat]['units']    += r['units'] or 0
        cats[cat]['revenue']  += r['revenue'] or 0
        cats[cat]['products'] += 1

    categories = [
        {'category': k, **v, 'avg_cvr': round(v['units']*100/max(v['units'],1), 1)}
        for k, v in sorted(cats.items(), key=lambda x: -x[1]['revenue'])
    ]

    return jsonify({
        'imports': imports,
        'top_performers': top[:50],
        'categories': categories,
        'totals': {
            'revenue': sum(r['revenue'] or 0 for r in all_rows),
            'units':   sum(r['units'] or 0 for r in all_rows),
        }
    })


@app.route('/api/sales/recommend', methods=['POST'])
def sales_recommend():
    """Use Claude to generate product recommendations based on sales data."""
    import anthropic

    top = SalesData.top_performers(limit=30, min_units=1)
    if not top:
        return jsonify({"error": "No sales data yet — import a report first"}), 400

    # Top 15 by revenue
    top_lines = "\n".join(
        f"- {r['sku']} | {r['title'][:70]} | {r['total_units']:.0f} units | "
        f"£{r['total_revenue']:.0f} revenue | {r['blended_cvr']:.1f}% CVR"
        for r in top[:15]
    )

    # Categories performing well
    all_rows = SalesData.category_summary()
    from collections import defaultdict
    cats = defaultdict(lambda: {'units': 0, 'revenue': 0, 'count': 0})
    for r in all_rows:
        cat = _infer_category(r['title'])
        cats[cat]['units']   += r['units'] or 0
        cats[cat]['revenue'] += r['revenue'] or 0
        cats[cat]['count']   += 1
    cat_lines = "\n".join(
        f"- {k}: {v['count']} products, {v['units']:.0f} units, £{v['revenue']:.0f} revenue"
        for k, v in sorted(cats.items(), key=lambda x: -x[1]['revenue'])[:10]
    )

    prompt = f"""You are a product development advisor for NBNE, a UK sign manufacturer selling aluminium composite signs on Amazon UK.

Our best-selling products from sales data:
{top_lines}

Performance by category:
{cat_lines}

Our signs are printed on brushed aluminium composite blanks in sizes:
- 9.5×9.5cm circular (dracula) — XS
- 11×9.5cm (saville) — S
- 14×9cm (dick) — M
- 19×14cm (barzan) — L
- 29×19cm with peel tab (baby_jesus) — XL

Based on what is ACTUALLY selling (high units, high CVR), recommend exactly 12 new sign products we should create next.
For each, provide:
1. Product title (as it would appear on Amazon UK)
2. Which blank size fits best
3. One sentence explaining why this will sell based on the data

Format each as:
PRODUCT: <title>
SIZE: <slug>
REASON: <why>

Focus on:
- Variants of proven winners (same category, different message or size)
- Underserved niches adjacent to our top categories
- High-volume UK search terms not yet covered
"""

    client = anthropic.Anthropic()
    message = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1500,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = message.content[0].text

    # Parse into structured list
    recommendations = []
    current = {}
    for line in raw.split('\n'):
        line = line.strip()
        if line.startswith('PRODUCT:'):
            if current.get('title'):
                recommendations.append(current)
            current = {'title': line[8:].strip(), 'size': '', 'reason': ''}
        elif line.startswith('SIZE:'):
            current['size'] = line[5:].strip()
        elif line.startswith('REASON:'):
            current['reason'] = line[7:].strip()
    if current.get('title'):
        recommendations.append(current)

    return jsonify({'recommendations': recommendations, 'raw': raw})


@app.route('/api/sales/imports')
def list_imports():
    return jsonify(SalesImport.list_all())


@app.route('/api/bug-report', methods=['POST'])
def bug_report():
    """Send a bug report email to the team."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from config import SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, BUG_REPORT_RECIPIENTS

    data = request.json or {}
    name = data.get('name', 'Unknown').strip() or 'Unknown'
    context = data.get('context', '').strip()
    description = data.get('description', '').strip()
    reporter_email = session.get('user_email', 'unknown')

    if not description:
        return jsonify({"error": "description required"}), 400

    body = f"""Bug report from Render ({reporter_email})

Reporter: {name}
While doing: {context or '(not specified)'}

--- Description ---
{description}
"""

    msg = MIMEMultipart()
    msg['From'] = SMTP_USER
    msg['To'] = ', '.join(BUG_REPORT_RECIPIENTS)
    msg['Subject'] = f'[Render] Bug report from {name}'
    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASSWORD)
            s.sendmail(SMTP_USER, BUG_REPORT_RECIPIENTS, msg.as_string())
        return jsonify({"ok": True})
    except Exception as e:
        app.logger.error("Bug report email failed: %s", e)
        return jsonify({"error": "mail failed"}), 500


# ── Auto-publish on QA approval ─────────────────────────────────────────────────

def _trigger_auto_publish(m_number: str):
    """Fire-and-forget publish to Etsy + Phloe shop when a product is QA-approved."""
    product = Product.get(m_number)
    if not product or product.get('qa_status') != 'approved':
        return

    # Load AI content
    content = None
    conn = get_db()
    try:
        cur = dict_cursor(conn)
        cur.execute(
            "SELECT title, description, bullet_points, search_terms "
            "FROM render_product_content WHERE product_id = %s "
            "ORDER BY created_at DESC LIMIT 1",
            (product['id'],)
        )
        row = cur.fetchone()
        if row:
            content = dict(row)
    finally:
        release_db(conn)

    content_map = {m_number: content} if content else {}

    # Publish to Etsy (background job)
    def _etsy_publish():
        try:
            from etsy_api import publish_products_to_etsy
            results = publish_products_to_etsy([product], content_map)
            _log_publish_results(results, 'etsy')
            app.logger.info("Auto-published %s to Etsy: %s", m_number,
                            results[0].get('status') if results else 'no result')
        except Exception as e:
            app.logger.error("Auto-publish %s to Etsy failed: %s", m_number, e)
            _log_publish_results([{
                'm_number': m_number, 'listing_id': None,
                'status': 'failed', 'error': str(e),
            }], 'etsy')

    # Publish to Phloe shop (background job)
    def _phloe_publish():
        try:
            from phloe_publisher import push_product_to_phloe
            result = push_product_to_phloe(product, content)
            _log_publish_results([{
                'm_number': m_number,
                'listing_id': result.get('product_id'),
                'status': result['status'],
                'error': result.get('error'),
            }], 'phloe')
            app.logger.info("Auto-published %s to Phloe: %s", m_number, result['status'])
        except Exception as e:
            app.logger.error("Auto-publish %s to Phloe failed: %s", m_number, e)
            _log_publish_results([{
                'm_number': m_number, 'listing_id': None,
                'status': 'failed', 'error': str(e),
            }], 'phloe')

    # Submit as background jobs so the PATCH response isn't blocked
    submit_job(f"Etsy publish {m_number}", _etsy_publish)
    submit_job(f"Phloe publish {m_number}", _phloe_publish)


# ── Etsy OAuth ──────────────────────────────────────────────────────────────────

@app.route('/etsy/oauth/connect')
def etsy_oauth_connect():
    """Start Etsy OAuth PKCE flow — redirects to Etsy consent page."""
    try:
        from etsy_auth import get_etsy_auth_from_env
        auth = get_etsy_auth_from_env()
        auth_url, code_verifier = auth.start_oauth_flow(state="etsy_render")
        session['etsy_code_verifier'] = code_verifier
        return redirect(auth_url)
    except ValueError as e:
        return jsonify({"error": str(e)}), 500


@app.route('/etsy/oauth/callback')
def etsy_oauth_callback():
    """Handle Etsy OAuth callback — exchange code for tokens."""
    from etsy_auth import get_etsy_auth_from_env

    code = request.args.get('code')
    error = request.args.get('error')
    if error:
        return jsonify({"error": error, "description": request.args.get('error_description')}), 400
    if not code:
        return jsonify({"error": "No authorization code received"}), 400

    code_verifier = session.pop('etsy_code_verifier', None)
    if not code_verifier:
        return jsonify({"error": "No code verifier in session — restart OAuth flow"}), 400

    try:
        auth = get_etsy_auth_from_env()
        auth.exchange_code_for_tokens(code, code_verifier)
        return redirect('/?etsy_connected=1')
    except Exception as e:
        app.logger.error("Etsy OAuth token exchange failed: %s", e)
        return jsonify({"error": f"Token exchange failed: {e}"}), 500


@app.route('/etsy/oauth/status')
def etsy_oauth_status():
    """Check Etsy OAuth connection status."""
    try:
        from etsy_auth import get_etsy_auth_from_env
        auth = get_etsy_auth_from_env()
        return jsonify(auth.get_status())
    except ValueError:
        return jsonify({"connected": False, "error": "Etsy API credentials not configured"})


# ── Etsy Publish ────────────────────────────────────────────────────────────────

@app.route('/api/etsy/publish', methods=['POST'])
def etsy_publish():
    """Publish QA-approved products to Etsy as draft listings.

    Body: { "m_numbers": ["M1001", "M1002"] } or { "all_approved": true }
    """
    from etsy_api import publish_products_to_etsy

    data = request.json or {}
    m_numbers = data.get('m_numbers', [])
    all_approved = data.get('all_approved', False)

    if all_approved:
        products = Product.approved()
    elif m_numbers:
        products = [p for p in [Product.get(m) for m in m_numbers] if p]
    else:
        return jsonify({"error": "Provide m_numbers list or all_approved: true"}), 400

    if not products:
        return jsonify({"error": "No products found"}), 404

    # Load AI content for each product
    content_map = {}
    conn = get_db()
    try:
        cur = dict_cursor(conn)
        for p in products:
            cur.execute(
                "SELECT title, description, bullet_points, search_terms "
                "FROM render_product_content WHERE product_id = %s "
                "ORDER BY created_at DESC LIMIT 1",
                (p['id'],)
            )
            row = cur.fetchone()
            if row:
                content_map[p['m_number']] = dict(row)
    finally:
        release_db(conn)

    dry_run = data.get('dry_run', False)
    results = publish_products_to_etsy(products, content_map, dry_run=dry_run)

    # Log publish results
    if not dry_run:
        _log_publish_results(results, 'etsy')

    published = sum(1 for r in results if r['status'] == 'success')
    failed = sum(1 for r in results if r['status'] == 'failed')

    return jsonify({
        "published": published,
        "failed": failed,
        "results": results,
    })


@app.route('/api/etsy/publish/status')
def etsy_publish_status():
    """Get recent Etsy publish log entries."""
    conn = get_db()
    try:
        cur = dict_cursor(conn)
        cur.execute(
            "SELECT * FROM render_publish_log WHERE channel = 'etsy' "
            "ORDER BY published_at DESC LIMIT 50"
        )
        return jsonify([dict(r) for r in cur.fetchall()])
    finally:
        release_db(conn)


@app.route('/api/phloe/publish', methods=['POST'])
def phloe_publish():
    """Publish QA-approved products to Phloe shop (app.nbnesigns.co.uk/shop).

    Body: { "m_numbers": ["M1001", "M1002"] } or { "all_approved": true }
    """
    from phloe_publisher import push_products_to_phloe

    data = request.json or {}
    m_numbers = data.get('m_numbers', [])
    all_approved = data.get('all_approved', False)

    if all_approved:
        products = Product.approved()
    elif m_numbers:
        products = [p for p in [Product.get(m) for m in m_numbers] if p]
    else:
        return jsonify({"error": "Provide m_numbers list or all_approved: true"}), 400

    if not products:
        return jsonify({"error": "No products found"}), 404

    # Load AI content
    content_map = {}
    conn = get_db()
    try:
        cur = dict_cursor(conn)
        for p in products:
            cur.execute(
                "SELECT title, description, bullet_points, search_terms "
                "FROM render_product_content WHERE product_id = %s "
                "ORDER BY created_at DESC LIMIT 1",
                (p['id'],)
            )
            row = cur.fetchone()
            if row:
                content_map[p['m_number']] = dict(row)
    finally:
        release_db(conn)

    results = push_products_to_phloe(products, content_map)

    # Log results
    for r in results:
        _log_publish_results([{
            'm_number': r['m_number'],
            'listing_id': r.get('product_id'),
            'status': r['status'],
            'error': r.get('error'),
        }], 'phloe')

    published = sum(1 for r in results if r['status'] == 'success')
    failed = sum(1 for r in results if r['status'] == 'failed')

    return jsonify({"published": published, "failed": failed, "results": results})


def _log_publish_results(results: list[dict], channel: str):
    """Write publish results to render_publish_log."""
    conn = get_db()
    try:
        cur = conn.cursor()
        for r in results:
            cur.execute(
                "INSERT INTO render_publish_log (m_number, channel, status, external_id, error_message) "
                "VALUES (%s, %s, %s, %s, %s)",
                (r['m_number'], channel, r['status'],
                 str(r.get('listing_id', '')) if r.get('listing_id') else None,
                 r.get('error')),
            )
    finally:
        release_db(conn)


# ── Cairn Context Endpoint ──────────────────────────────────────────────────────

@app.route('/api/cairn/context')
def cairn_context():
    """Expose Render state to Cairn business brain."""
    conn = get_db()
    try:
        cur = dict_cursor(conn)

        cur.execute("SELECT COUNT(*) as n FROM render_products")
        total = dict(cur.fetchone())['n']

        cur.execute("SELECT COUNT(*) as n FROM render_products WHERE qa_status = 'approved'")
        approved = dict(cur.fetchone())['n']

        cur.execute("SELECT COUNT(*) as n FROM render_products WHERE qa_status = 'pending'")
        pending = dict(cur.fetchone())['n']

        cur.execute(
            "SELECT channel, COUNT(*) as n FROM render_publish_log "
            "WHERE status = 'success' GROUP BY channel"
        )
        publish_counts = {dict(r)['channel']: dict(r)['n'] for r in cur.fetchall()}

        cur.execute(
            "SELECT m_number, channel, status, published_at FROM render_publish_log "
            "ORDER BY published_at DESC LIMIT 10"
        )
        recent = [dict(r) for r in cur.fetchall()]
        for r in recent:
            if hasattr(r.get('published_at'), 'isoformat'):
                r['published_at'] = r['published_at'].isoformat()

        return jsonify({
            "module": "render",
            "generated_at": datetime.utcnow().isoformat(),
            "products": {
                "total": total,
                "approved": approved,
                "pending_qa": pending,
            },
            "publishing": publish_counts,
            "recent_activity": recent,
            "summary": (
                f"{total} products ({approved} approved, {pending} pending QA). "
                f"Published: {publish_counts}."
            ),
        })
    finally:
        release_db(conn)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
